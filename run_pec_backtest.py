#!/usr/bin/env python3
# run_pec_backtest.py
# Standalone Post-Entry Control (PEC) backtesting script
# Run this independently from main.py to test signal quality
# Usage: python3 run_pec_backtest.py

print("[PEC_BACKTEST] Script started.", flush=True)

from signal_debug_log import export_signal_debug_txt, log_fired_signal
import os
import time
import pandas as pd
import pytz
from datetime import datetime, timedelta
from kucoin_data import get_ohlcv, DEFAULT_SLIPPAGE
from smart_filter import SmartFilter
from telegram_alert import send_telegram_alert, send_telegram_file
from pec_backtest import load_fired_signals_from_log, find_closest_ohlcv_bar
from pec_engine import run_pec_check, export_pec_log
from tp_sl_retracement import calculate_tp_sl
import math

WIB = pytz.timezone('Asia/Jakarta')

# --- Configuration ---
TOKENS = [
    "BTC-USDT", "ETH-USDT",
    "BNB-USDT", "XRP-USDT", "SOL-USDT", "ADA-USDT", "XLM-USDT",
    "TON-USDT", "AVAX-USDT", "LINK-USDT", "DOT-USDT", "ARB-USDT",
    "PUMP-USDT", "KAITO-USDT", "MAGIC-USDT", "SUI-USDT", "AERO-USDT",
    "BERA-USDT", "UNI-USDT", "HBAR-USDT", "SAHARA-USDT", "VIRTUAL-USDT",
    "PARTI-USDT", "CFX-USDT", "DOGE-USDT", "VINE-USDT", "PENGU-USDT",
    "WIF-USDT", "EIGEN-USDT", "SPK-USDT", "HYPE-USDT", "WLFI-USDT",
    "POL-USDT", "RAY-USDT", "ZKJ-USDT", "AAVE-USDT", "DYDX-USDT",
    "ONDO-USDT", "ARKM-USDT", "ATH-USDT", "NMR-USDT", "PROMPT-USDT",
    "TURBO-USDT", "ENA-USDT", "BIO-USDT", "ASTER-USDT", "XPL-USDT",
    "AVNT-USDT", "ORDER-USDT", "XAUT-USDT", "ZORA-USDT"
]

PEC_BARS = 5
PEC_WINDOW_MINUTES = 720
OHLCV_LIMIT = 250

def get_local_wib(dt):
    """Convert datetime to WIB (Asia/Jakarta) timezone"""
    if not isinstance(dt, pd.Timestamp):
        dt = pd.Timestamp(dt)
    return dt.tz_localize('UTC').tz_convert(WIB).replace(microsecond=0).strftime('%Y-%m-%d %H:%M:%S')

def run_pec_backtest():
    """
    Standalone PEC backtesting function.
    Loads fired signals from logs, fetches OHLCV data, and runs PEC analysis.
    """
    print("[PEC_BACKTEST] Starting PEC backtest...", flush=True)
    
    # Load fired signals from logs
    print(f"[PEC_BACKTEST] Loading fired signals from logs (last {PEC_WINDOW_MINUTES} minutes)...", flush=True)
    signals, unique_symbols, unique_timeframes = load_fired_signals_from_log(
        log_file_path="logs.txt",
        minutes_limit=PEC_WINDOW_MINUTES
    )
    
    if not signals:
        print("[PEC_BACKTEST] No fired signals found in logs. Exiting.", flush=True)
        return
    
    print(f"[PEC_BACKTEST] Loaded {len(signals)} signals", flush=True)
    print(f"[PEC_BACKTEST] Unique symbols: {sorted(unique_symbols)}", flush=True)
    print(f"[PEC_BACKTEST] Unique timeframes: {sorted(unique_timeframes)}", flush=True)
    
    # Process each signal with PEC
    pec_results = []
    
    for idx, signal in enumerate(signals, start=1):
        symbol = signal.get("symbol", "").strip()
        tf = signal.get("tf", "").strip()
        signal_type = signal.get("signal_type", "").strip()
        entry_price = float(signal.get("entry_price", 0.0))
        entry_idx_from_log = int(signal.get("entry_idx", 0))
        fired_time_str = signal.get("entry_time", "")
        
        try:
            # Parse fired time
            if 'T' in fired_time_str:
                fired_time_utc = pd.Timestamp(fired_time_str)
            else:
                fired_time_utc = pd.Timestamp(fired_time_str)
        except Exception as e:
            print(f"[PEC_BACKTEST] Error parsing time for signal {idx}: {e}", flush=True)
            continue
        
        try:
            # Fetch OHLCV data
            print(f"[PEC_BACKTEST] [{idx}/{len(signals)}] {symbol} {tf} {signal_type} @ ${entry_price:.5f}...", flush=True)
            df = get_ohlcv(symbol, interval=tf, limit=OHLCV_LIMIT)
            
            if df is None or df.empty or len(df) < PEC_BARS + 5:
                print(f"[PEC_BACKTEST]     ⚠️ Insufficient data for PEC. Skipping.", flush=True)
                continue
            
            # Run PEC check
            score = signal.get("score", 0)
            confidence = signal.get("confidence", 0)
            passed_gk = signal.get("passed_gk", 0)
            max_gk = signal.get("max_gk", 0)
            
            pec_result = run_pec_check(
                symbol=symbol,
                fired_time_utc=fired_time_utc,
                tf=tf,
                signal_type=signal_type,
                entry_price=entry_price,
                ohlcv_df=df,
                pec_bars=PEC_BARS,
                score=score,
                confidence=confidence,
                passed_gk_count=passed_gk
            )
            
            if pec_result and "status" not in pec_result:
                # Valid PEC result
                mfe = pec_result.get("max_favorable", 0.0)
                mae = pec_result.get("max_adverse", 0.0)
                final_ret = pec_result.get("final_return", 0.0)
                
                # Determine outcome
                outcome = "✅ WIN" if final_ret > 0.5 else ("❌ LOSS" if final_ret < -0.5 else "➖ BREAK")
                
                print(f"[PEC_BACKTEST]     {outcome} | MFE: {mfe:+.2f}% | MAE: {mae:+.2f}% | Return: {final_ret:+.2f}%", flush=True)
                
                pec_results.append({
                    "symbol": symbol,
                    "tf": tf,
                    "signal_type": signal_type,
                    "entry_price": entry_price,
                    "confidence": confidence,
                    "score": score,
                    "mfe": mfe,
                    "mae": mae,
                    "final_return": final_ret,
                    "outcome": outcome,
                    "pec_summary": pec_result.get("summary", "")
                })
                
                # Export PEC log for this signal
                try:
                    export_pec_log(pec_result, filename="pec_backtest_results.txt")
                except Exception as e:
                    print(f"[PEC_BACKTEST]     ⚠️ Error exporting PEC log: {e}", flush=True)
            else:
                print(f"[PEC_BACKTEST]     ⚠️ PEC check failed. Skipping.", flush=True)
        
        except Exception as e:
            print(f"[PEC_BACKTEST]     ❌ Exception: {e}", flush=True)
            import traceback
            traceback.print_exc()
    
    # Generate summary report
    if pec_results:
        print(f"\n{'='*80}", flush=True)
        print(f"[PEC_BACKTEST] BACKTEST SUMMARY", flush=True)
        print(f"{'='*80}", flush=True)
        
        df_results = pd.DataFrame(pec_results)
        
        total_signals = len(df_results)
        winning = len(df_results[df_results["final_return"] > 0.5])
        losing = len(df_results[df_results["final_return"] < -0.5])
        breakeven = total_signals - winning - losing
        
        win_rate = (winning / total_signals * 100) if total_signals > 0 else 0
        avg_win = df_results[df_results["final_return"] > 0.5]["final_return"].mean() if winning > 0 else 0
        avg_loss = df_results[df_results["final_return"] < -0.5]["final_return"].mean() if losing > 0 else 0
        avg_mfe = df_results["mfe"].mean()
        avg_mae = df_results["mae"].mean()
        
        print(f"\nTotal Signals Tested: {total_signals}", flush=True)
        print(f"Winning: {winning} ({win_rate:.1f}%)", flush=True)
        print(f"Losing:  {losing} ({(losing/total_signals*100):.1f}%)", flush=True)
        print(f"Break-Even: {breakeven}", flush=True)
        print(f"\nAverage Win: {avg_win:+.2f}%", flush=True)
        print(f"Average Loss: {avg_loss:+.2f}%", flush=True)
        print(f"Average MFE: {avg_mfe:+.2f}%", flush=True)
        print(f"Average MAE: {avg_mae:+.2f}%", flush=True)
        
        print(f"\n{'='*80}", flush=True)
        print(f"[PEC_BACKTEST] Per-Symbol Breakdown:", flush=True)
        print(f"{'='*80}", flush=True)
        
        by_symbol = df_results.groupby("symbol").agg({
            "final_return": ["count", "mean"],
            "confidence": "mean",
            "score": "mean"
        }).round(2)
        
        print(by_symbol.to_string())
        
        print(f"\n{'='*80}", flush=True)
        print(f"[PEC_BACKTEST] Per-Timeframe Breakdown:", flush=True)
        print(f"{'='*80}", flush=True)
        
        by_tf = df_results.groupby("tf").agg({
            "final_return": ["count", "mean"],
            "confidence": "mean",
            "score": "mean"
        }).round(2)
        
        print(by_tf.to_string())
        
        # Save results to both CSV and XLSX
        timestamp = datetime.now(WIB).strftime('%Y%m%d_%H%M%S')
        
        # CSV export
        try:
            csv_path = f"pec_backtest_results_{timestamp}.csv"
            df_results.to_csv(csv_path, index=False)
            print(f"\n✅ CSV saved to: {csv_path}", flush=True)
        except Exception as e:
            print(f"⚠️ Error saving CSV: {e}", flush=True)
        
        # XLSX export with formatting
        try:
            xlsx_path = f"pec_backtest_results_{timestamp}.xlsx"
            
            # Create Excel writer
            with pd.ExcelWriter(xlsx_path, engine='openpyxl') as writer:
                df_results.to_excel(writer, sheet_name='PEC Results', index=False)
                
                # Get the workbook and worksheet
                workbook = writer.book
                worksheet = writer.sheets['PEC Results']
                
                # Add formatting
                from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                
                # Header formatting
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Auto-width columns
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Color rows by outcome
                win_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                loss_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red
                breakeven_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                
                # Find outcome column
                outcome_col = None
                for idx, cell in enumerate(worksheet[1], 1):
                    if cell.value == "outcome":
                        outcome_col = idx
                        break
                
                if outcome_col:
                    for row in worksheet.iter_rows(min_row=2, max_row=len(df_results) + 1):
                        outcome = row[outcome_col - 1].value
                        if "WIN" in str(outcome):
                            for cell in row:
                                cell.fill = win_fill
                        elif "LOSS" in str(outcome):
                            for cell in row:
                                cell.fill = loss_fill
                        elif "BREAK" in str(outcome):
                            for cell in row:
                                cell.fill = breakeven_fill
            
            print(f"✅ Excel saved to: {xlsx_path}", flush=True)
        except ImportError:
            print(f"⚠️ openpyxl not installed. Install with: pip install openpyxl", flush=True)
        except Exception as e:
            print(f"⚠️ Error saving XLSX: {e}", flush=True)
    else:
        print("[PEC_BACKTEST] No valid PEC results generated.", flush=True)
    
    print(f"\n[PEC_BACKTEST] Backtest complete.", flush=True)

if __name__ == "__main__":
    print("[PEC_BACKTEST] ========== STANDALONE PEC BACKTEST ==========", flush=True)
    print(f"[PEC_BACKTEST] Time: {get_local_wib(datetime.utcnow())}", flush=True)
    print("[PEC_BACKTEST] This script tests signal quality using realistic exit logic", flush=True)
    print("[PEC_BACKTEST] (TP: +1.5%, SL: -1.0%, Fee: 0.1%)", flush=True)
    print("[PEC_BACKTEST] ==========================================\n", flush=True)
    
    try:
        run_pec_backtest()
    except Exception as e:
        print(f"[PEC_BACKTEST] Fatal error: {e}", flush=True)
        import traceback
        traceback.print_exc()
